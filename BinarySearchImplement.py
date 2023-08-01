import openpyxl as op
v1 = op.load_workbook("Excel.xlsx")
v2 = v1.active
main_lis = []
for i in range(1, v2.max_row):
    for j in v2.iter_cols(2, 2):
        main_lis.append(j[i].value)
for i in range(len(main_lis)):
    for j in range(i + 1, len(main_lis)):
        if main_lis[i] > main_lis[j]:
            main_lis[i], main_lis[j] = main_lis[j], main_lis[i]
print(main_lis)


def BinarySearch(arr, x, low, high):
    while low <= high:
        mid = (low + high) // 2
        if arr[mid] == x:
            return mid
        elif arr[mid] < x:
            low = mid + 1
        else:
            high = mid - 1
    return None


x1 = input("Enter Name: ")
s = BinarySearch(main_lis, x1, 0, len(main_lis) - 1)
if s is None:
    print("Not Found")
else:
    print("Name Found At Index: ", s)

